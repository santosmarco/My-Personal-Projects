import openpyxl, shelve

users_file = shelve.open('users')

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'crpytobot\'s users'

ws['A1'] = 'ID'
ws['B1'] = 'Name'
ws['C1'] = 'Registered since'
ws['D1'] = 'Is auto_news on?'
ws['E1'] = 'Portfolio'

for i in range(5):
    ws.column_dimensions[openpyxl.cell.get_column_letter(i+1)].width = 20

row = 2
for user in users_file:
    ws.cell(row=row, column=1).value = user
    user = users_file[user]
    ws.cell(row=row, column=2).value = user['f_name']+' '+user['l_name']
    ws.cell(row=row, column=3).value = user['date']
    ws.cell(row=row, column=4).value = 'Y' if user['auto_news'] else 'N'
    portfolio_str = ''
    for coin in user['portfolio']:
        portfolio_str += '{}: {}\n'.format(coin, user['portfolio']['coin'])
    ws.cell(row=row, column=5).value = portfolio_str
    row += 1

wb.save('users.xlsx')
users_file.close()
